import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from bs4 import BeautifulSoup
from selenium.common.exceptions import TimeoutException
from datetime import datetime

today_date = datetime.now().strftime('%Y-%m-%d')

options = webdriver.ChromeOptions()
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
driver = webdriver.Chrome(options=options)
driver.get(f"https://www.target.com/c/tvs-home-theater-electronics/all-deals/-/N-5xtdwZakkos")
time.sleep(10)

products_data = []


for i in range(1, 32):  
    try:                                               
        product_name = driver.find_element    (By.XPATH, f'/html/body/div[1]/div[2]/main/div/div[1]/div/div/div[3]/div/div/div[1]/section/div/div[{i}]/div/div/div/div[2]/div/div/div[2]/div[2]/a').text

        price = driver.find_element           (By.XPATH, f'/html/body/div[1]/div[2]/main/div/div[1]/div/div/div[3]/div/div/div[1]/section/div/div[{i}]/div/div/div/div[2]/div/div/div[1]/div/div[1]/span').text
        
        product_details = driver.find_element(By.XPATH,  f'/html/body/div[1]/div[2]/main/div/div[1]/div/div/div[3]/div/div/div[1]/section/div/div[{i}]/div/div/div/div[2]/div/div/div[2]/div[1]/div/a').text

        Reviews = driver.find_element        (By.XPATH,  f'/html/body/div[1]/div[2]/main/div/div[1]/div/div/div[3]/div/div/div[1]/section/div/div[{i}]/div/div/div/div[2]/div/div/div[2]/a/div/span[2]').text

        products_data.append({"Product Name": product_name, 
                              "Price": price ,
                              "product details": product_details, 
                              "Reviews": Reviews,
                              "Date": today_date})
        
        if i % 4 == 0:
            driver.execute_script("window.scrollBy(0, 800);")  # Scrolls down by 1000 pixels
            time.sleep(2)  # Add a short delay to allow content to load
    except Exception as e:
        print(f"Error on item {i}: {e}")

driver.quit()

output_file = r"G:\Work\Scraping\sciencedirect\scraped_products.xlsx"

# Convert scraped data to a DataFrame
df = pd.DataFrame(products_data)

try:
    # Try to load the existing workbook
    workbook = load_workbook(output_file)
    
    # Load the writer and append data to the existing sheet
    with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        # Write data below existing data
        df.to_excel(writer, index=False, header=False, startrow=writer.sheets["Sheet1"].max_row)
    print(f"Data successfully appended to {output_file}")
except FileNotFoundError:
    # If file does not exist, create a new one
    df.to_excel(output_file, index=False)
    print(f"New file created and data saved to {output_file}")

#/html/body/div[1]/div[2]/main/div/div[1]/div/div/div[3]/div/div/div[1]/section/div/div[1]
#/html/body/div[1]/div[2]/main/div/div[1]/div/div/div[3]/div/div/div[1]/section/div/div[2]
#/html/body/div[1]/div[2]/main/div/div[1]/div/div/div[3]/div/div/div[1]/section/div/div[31]

 # /html/body/div[1]/div[2]/main/div/div[1]/div/div/div[3]/div/div/div[1]/section/div/div[i]/div/div/div/div[2]/div/div/div[1]/div/div[1]/span
